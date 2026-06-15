#!/usr/bin/env python3
"""
Extract Oasis Payer360 Data Dictionary (Excel) into RAG markdown documents.

Reads all 22 sheets and generates:
  data/rag/payer360/payer360_glossary.md
  data/rag/payer360/payer360_faq.md
  data/rag/payer360/payer360_things_to_know.md
  data/rag/payer360/payer360_tables.md
  data/rag/payer360/payer360_business_rules.md
  data/catalog/table_catalog.json
  data/catalog/column_catalog.json
  data/catalog/relationship_catalog.json
  data/rag/payer360/payer360_catalog.md
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "Oasis Payer360 - Data Dictionary.xlsx"
RAG_DIR = ROOT / "data" / "rag" / "payer360"
CATALOG_DIR = ROOT / "data" / "catalog"
RAG_DIR.mkdir(parents=True, exist_ok=True)
CATALOG_DIR.mkdir(parents=True, exist_ok=True)

# Fix Windows console encoding
import io as _io
sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── Sheet → Subject area mapping ───────────────────────────────────────────────
SUBJECT_AREAS = {
    "Oasis Payer360 - Summary": "overview",
    "MCO Profile & Hierarchy - Norma": "mco",
    "Formulary and Access": "formulary",
    "LAAD Claims - Base Fact & Aggre": "claims",
    "Sales - Base Fact & Aggregate": "sales",
    "Sales indication- Base Fact & A": "sales",
    "MCO Lives - Base Fact and Aggre": "lives",
    "Provider Payer Mix - Summarized": "provider_payer_mix",
    "GPO Claims - Norm & GPO Claims ": "gpo_claims",
    "PSU Claims - Norm & PSU Aggrega": "psu_claims",
    "MCO-HCP Rollup": "provider_payer_mix",
    "Formulary ProductPayerHPMPEN St": "formulary",
    "Formulary Payer Lives": "lives",
    "Plantrak Sales w WAC dollar val": "sales",
    "Annual Book Of Business": "mco",
    "Precision AQ": "precision_aq",
    "Rollups Delta Tables": "formulary",
}


def safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def is_col_row(row: tuple) -> bool:
    """Detect column definition rows: has table name, column name, data type."""
    if len(row) < 5:
        return False
    table = safe_str(row[1])
    col = safe_str(row[2])
    dtype = safe_str(row[4])
    return (
        bool(table)
        and bool(col)
        and col not in ("Column Name", "Table and Attributes", "Oasis Table Name")
        and bool(dtype)
        and dtype in ("string", "int", "bigint", "float", "date", "boolean", "timestamp")
    )


def extract_workbook(wb: openpyxl.Workbook) -> dict:
    """Parse all sheets into structured data."""
    data = {
        "tables": [],       # {name, schema, description, subject_area, sample_questions}
        "columns": [],      # {table, column, data_type, description, is_key, subject_area}
        "things_to_know": [],  # {subject_area, text}
        "sample_questions": [],  # {subject_area, question, table}
        "business_rules": [],   # {subject_area, rule}
    }

    seen_tables: set[str] = set()

    for sheet_name in wb.sheetnames:
        subject = SUBJECT_AREAS.get(sheet_name, "general")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        current_table = None
        current_schema = None
        things_to_know_text = ""
        sample_q_text = ""

        for row in rows:
            if not any(c for c in row if c is not None):
                continue

            r = [safe_str(c) for c in row]

            # Detect "Things to Know" column (col index 5)
            if len(row) > 5 and row[5] and "Things to know" not in safe_str(row[5]):
                ttk = safe_str(row[5])
                if len(ttk) > 20:
                    things_to_know_text = ttk
                    # Split numbered items
                    for line in ttk.split("\\n"):
                        line = line.strip()
                        if line and len(line) > 10:
                            data["things_to_know"].append({
                                "subject_area": subject,
                                "sheet": sheet_name,
                                "text": line,
                            })

            # Detect sample questions (col index 2 on table header rows)
            if len(row) > 2 and row[2] and "?" in safe_str(row[2]) and len(safe_str(row[2])) > 20:
                q = safe_str(row[2])
                # Find associated table
                tbl = safe_str(row[1]) if row[1] else current_table or ""
                data["sample_questions"].append({
                    "subject_area": subject,
                    "sheet": sheet_name,
                    "question": q,
                    "table": tbl,
                })

            # Detect table name rows — pattern: col1=None, col1=table_path
            if len(row) > 1 and row[1] and "payer_360_" in safe_str(row[1]):
                raw = safe_str(row[1])
                # Could be "oasis_normalized.payer_360_xxx (Normalized)" or just table name
                if "." in raw:
                    parts = raw.split(".", 1)
                    schema = parts[0].strip()
                    tbl_raw = parts[1].split("(")[0].split("\n")[0].strip()
                elif "(" in raw:
                    schema = "oasis_normalized"
                    tbl_raw = raw.split("(")[0].strip()
                else:
                    schema = "oasis_normalized"
                    tbl_raw = raw.strip()

                tbl_raw = tbl_raw.strip()
                if tbl_raw and tbl_raw not in seen_tables:
                    seen_tables.add(tbl_raw)
                    # Get description from col5 (Things to Know col) if present
                    desc = ""
                    if len(row) > 5 and row[5]:
                        desc = safe_str(row[5])[:200]
                    # Get sample q from col2
                    sample_q = safe_str(row[2]) if len(row) > 2 else ""
                    data["tables"].append({
                        "name": tbl_raw,
                        "schema": schema,
                        "subject_area": subject,
                        "sheet": sheet_name,
                        "description": desc,
                        "sample_question": sample_q,
                    })
                current_table = tbl_raw
                current_schema = schema

            # Detect column definition rows
            if is_col_row(row):
                col_name = safe_str(row[2])
                dtype = safe_str(row[4])
                desc = safe_str(row[5]) if len(row) > 5 else ""
                is_key = "PK" in safe_str(row[3]) or "CPK" in safe_str(row[3]) or "FK" in safe_str(row[3])
                key_type = safe_str(row[3]) if row[3] else ""

                # Use table from col1 if present, else current_table
                tbl = safe_str(row[1]) if row[1] and "payer_360_" in safe_str(row[1]) else (current_table or "")
                if "." in tbl:
                    tbl = tbl.split(".", 1)[1].split("(")[0].strip()

                if col_name and tbl:
                    data["columns"].append({
                        "table": tbl,
                        "schema": current_schema or "oasis_normalized",
                        "column": col_name,
                        "data_type": dtype,
                        "description": desc[:300],
                        "is_key": is_key,
                        "key_type": key_type,
                        "subject_area": subject,
                    })

    return data


def write_glossary(data: dict):
    """payer360_glossary.md — all column definitions grouped by table."""
    lines = [
        "# Payer360 Data Dictionary — Glossary\n",
        "All column definitions extracted from the Oasis Payer360 Data Dictionary.\n",
        "Source: Oasis Payer360 - Data Dictionary.xlsx\n\n",
    ]

    # Group by table
    by_table: dict[str, list] = {}
    for col in data["columns"]:
        by_table.setdefault(col["table"], []).append(col)

    for table, cols in sorted(by_table.items()):
        schema = cols[0]["schema"]
        subj = cols[0]["subject_area"]
        lines.append(f"## Table: {table}\n")
        lines.append(f"Schema: {schema} | Subject Area: {subj}\n\n")
        for col in cols:
            key_flag = f" [{col['key_type']}]" if col["is_key"] else ""
            desc = col["description"] or "No description available."
            lines.append(f"- **{col['column']}** ({col['data_type']}){key_flag}: {desc}\n")
        lines.append("\n")

    (RAG_DIR / "payer360_glossary.md").write_text("".join(lines), encoding="utf-8")
    print(f"  ✓ payer360_glossary.md ({len(data['columns'])} columns, {len(by_table)} tables)")


def write_faq(data: dict):
    """payer360_faq.md — all sample questions with table mappings."""
    lines = [
        "# Payer360 — Sample Business Questions\n\n",
        "These questions can be answered using the Payer360 data platform.\n\n",
    ]

    by_subject: dict[str, list] = {}
    for q in data["sample_questions"]:
        by_subject.setdefault(q["subject_area"], []).append(q)

    for subj, questions in sorted(by_subject.items()):
        lines.append(f"## {subj.replace('_', ' ').title()}\n\n")
        seen_qs: set[str] = set()
        for q in questions:
            qtext = q["question"]
            if qtext not in seen_qs and len(qtext) > 10:
                seen_qs.add(qtext)
                tbl = q["table"].split("\n")[0].strip() if q["table"] else ""
                if tbl:
                    lines.append(f"- {qtext}\n  - *Source table: {tbl}*\n")
                else:
                    lines.append(f"- {qtext}\n")
        lines.append("\n")

    (RAG_DIR / "payer360_faq.md").write_text("".join(lines), encoding="utf-8")
    print(f"  ✓ payer360_faq.md ({sum(len(v) for v in by_subject.values())} Q&As)")


def write_things_to_know(data: dict):
    """payer360_things_to_know.md — caveats and gotchas."""
    lines = [
        "# Payer360 — Things To Know (Business Rules & Caveats)\n\n",
        "Critical business rules from the Oasis Payer360 Data Dictionary.\n",
        "These rules govern correct usage of the data and must be applied.\n\n",
    ]

    by_subject: dict[str, list] = {}
    for item in data["things_to_know"]:
        by_subject.setdefault(item["subject_area"], []).append(item)

    for subj, items in sorted(by_subject.items()):
        lines.append(f"## {subj.replace('_', ' ').title()}\n\n")
        seen: set[str] = set()
        for item in items:
            txt = item["text"]
            # Deduplicate and filter noise
            if txt not in seen and len(txt) > 15 and txt not in ("Things to know:", "nan"):
                seen.add(txt)
                lines.append(f"- {txt}\n")
        lines.append("\n")

    (RAG_DIR / "payer360_things_to_know.md").write_text("".join(lines), encoding="utf-8")
    print(f"  ✓ payer360_things_to_know.md ({len(data['things_to_know'])} items)")


def write_tables(data: dict):
    """payer360_tables.md — all table descriptions."""
    lines = [
        "# Payer360 — Table Reference\n\n",
        "All tables in the Oasis Payer360 platform.\n\n",
        "| Table | Schema | Subject Area | Description |\n",
        "|---|---|---|---|\n",
    ]

    for tbl in sorted(data["tables"], key=lambda x: x["name"]):
        desc = tbl["description"].replace("|", "/").replace("\n", " ")[:120]
        lines.append(f"| {tbl['name']} | {tbl['schema']} | {tbl['subject_area']} | {desc} |\n")

    lines.append("\n\n## Tables by Subject Area\n\n")
    by_subj: dict[str, list] = {}
    for t in data["tables"]:
        by_subj.setdefault(t["subject_area"], []).append(t)

    for subj, tables in sorted(by_subj.items()):
        lines.append(f"### {subj.replace('_', ' ').title()}\n\n")
        for t in tables:
            sample = t.get("sample_question", "").split("\n")[0][:100]
            lines.append(f"- **{t['schema']}.{t['name']}**\n")
            if t["description"]:
                lines.append(f"  - {t['description'][:150]}\n")
            if sample:
                lines.append(f"  - Example question: _{sample}_\n")
        lines.append("\n")

    (RAG_DIR / "payer360_tables.md").write_text("".join(lines), encoding="utf-8")
    print(f"  ✓ payer360_tables.md ({len(data['tables'])} tables)")


def write_business_rules(data: dict):
    """payer360_business_rules.md — synthesized business rules."""
    content = """# Payer360 — Business Rules

## Sales Data Rules
- Plantrak and Imputed Sales are NOT additive. Always use the `sales_source` attribute to filter.
- Imputed Dollar Sales is available for GNE products only.
- Plantrak is IQVIA-sourced pharmacy sales data for HCPs. Units and Dollars are available.
- All Sales data is available at Product Level only. No Indication data is available in the sales table.

## LAAD Claims Rules
- LAAD covers 8 TAs: AHR, RESP, RA, MS, 5GLT, OPTHA, HCC, LN.
- Market Share Metric varies by TA: MS=Patient Count, AHR=Units, OPTHA=Claim Count, RA=Patient count (do not combine benefit types), RESP=Patient Count.
- Use `flag_market_share_eligible` flag in base fact when conducting analytics.
- Sum(Total Patient Count) in a quarter does NOT equal Count(Distinct Patient) in the same quarter. Be careful when rolling up time periods.
- In aggregates, `patient_type` has values: NEW, CONTINUING, ALL. Apply relevant filters to avoid double counting.
- `ecosystem_name` in aggregates includes NATIONAL level roll-ups. Apply relevant filters when using metrics.

## MCO Hierarchy Rules
- MCO entities have parent-child relationships maintained in `payer_360_mco_hierarchy`.
- `payer_360_mco_hierarchy_benefittype` combines hierarchy with benefit type from transaction data.
- Parent MCO entities may span multiple book of business categories.

## Formulary Rules
- Formulary status uses Health Plan Management (HPM) values and rankings.
- Has 3-year history.
- Access position compares GNE brand vs. competitors: Advantage, Parity, or Disadvantage.
- Override table (`payer_360_formulary_override`) captures impact team DCRs.
- Rollup table (`payer_360_formulary_rollup`) provides consolidated HPM values at rolled-up payer levels.

## Lives Data Rules
- Zip level lives data sourced from DRG.
- Two grains: (1) Plan-Payer: Pharmacy and Medical Benefit Lives, (2) Payer-PBM: Pharmacy Benefit Lives.
- Has 3-year history.

## GPO Portal Rules
- No Patient ID in GPO Portal data — unique patient counts cannot be derived.
- Only ASCENT provides enrollment data (lives enrolled to formulary).
- Data is NOT at Claim level.

## Provider Payer Mix Rules
- Available at HCP-Payer grain only.
- LAAD Metrics should be used for the 7 prioritized TAs: MS, OPTHA, RESP, IMM, AHR, 5GLT, HCC.
- GPO and PSU metrics for other TAs.
- Plantrak Sales Metrics when looking across TAs and indications.
- Data available at 2 grains specified by `aggregate_level` column.

## Payer Submitted Claims (PSD) Rules
- Patient ID in PSD data is NOT longitudinal across payers. Unique patient counts cannot be derived.
- Data may have time lag — payer submits only when rebate is due (usually quarter end + 90 days).

## Schema Rules
- Raw / normalized data lives in: `oasis_normalized` schema.
- Aggregated / summarized data lives in: `oasis_summarized` schema.
- Always prefer `oasis_summarized` tables for analytics unless raw claim detail is needed.
"""
    (RAG_DIR / "payer360_business_rules.md").write_text(content, encoding="utf-8")
    print("  ✓ payer360_business_rules.md")


def write_catalog(data: dict):
    """Write JSON catalogs + RAG-friendly catalog.md."""
    # table_catalog.json
    (CATALOG_DIR / "table_catalog.json").write_text(
        json.dumps(data["tables"], indent=2, ensure_ascii=False), encoding="utf-8"
    )

    # column_catalog.json
    (CATALOG_DIR / "column_catalog.json").write_text(
        json.dumps(data["columns"], indent=2, ensure_ascii=False), encoding="utf-8"
    )

    # relationship_catalog.json — inferred from key columns
    relationships = []
    key_cols = [c for c in data["columns"] if c["is_key"] and c["key_type"] in ("FK", "CPK")]
    join_patterns = [
        ("mdm_mco_id", "payer_360_mco_profile", "mdm_mco_id"),
        ("hcp_mdm_id", "payer_360_provider_payer_mix", "provider_mdm_id"),
        ("oasis_product_brand_id", "payer_360_sales", "oasis_product_brand_id"),
    ]
    for col in key_cols:
        for pattern_col, target_table, target_col in join_patterns:
            if col["column"] == pattern_col and col["table"] != target_table:
                relationships.append({
                    "from_table": col["table"],
                    "from_column": col["column"],
                    "to_table": target_table,
                    "to_column": target_col,
                    "relationship_type": "many_to_one",
                })

    (CATALOG_DIR / "relationship_catalog.json").write_text(
        json.dumps(relationships, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    # payer360_catalog.md — text version for RAG
    lines = [
        "# Payer360 AI Data Catalog\n\n",
        "Use this to answer questions like: which table contains X, where is Y stored.\n\n",
    ]
    lines.append("## Quick Reference: Table → Subject Area → Schema\n\n")
    for tbl in sorted(data["tables"], key=lambda x: x["subject_area"]):
        lines.append(f"- `{tbl['schema']}.{tbl['name']}` — {tbl['subject_area'].replace('_',' ').title()}\n")

    lines.append("\n## Column Search Index\n\n")
    lines.append("Key columns and which tables contain them:\n\n")
    col_index: dict[str, list[str]] = {}
    for col in data["columns"]:
        col_index.setdefault(col["column"], []).append(col["table"])
    for col_name, tables in sorted(col_index.items()):
        if len(tables) > 1:
            lines.append(f"- **{col_name}**: {', '.join(tables)}\n")

    (RAG_DIR / "payer360_catalog.md").write_text("".join(lines), encoding="utf-8")

    print(f"  ✓ table_catalog.json ({len(data['tables'])} tables)")
    print(f"  ✓ column_catalog.json ({len(data['columns'])} columns)")
    print(f"  ✓ relationship_catalog.json ({len(relationships)} relationships)")
    print("  ✓ payer360_catalog.md")


def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading: {EXCEL_PATH.name}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    print(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames[:5]}...")

    print("\nExtracting data from all sheets...")
    data = extract_workbook(wb)
    wb.close()

    print(f"\nExtracted: {len(data['tables'])} tables, {len(data['columns'])} columns, "
          f"{len(data['sample_questions'])} Q&As, {len(data['things_to_know'])} caveats")

    print("\nWriting RAG documents...")
    write_glossary(data)
    write_faq(data)
    write_things_to_know(data)
    write_tables(data)
    write_business_rules(data)

    print("\nWriting catalog files...")
    write_catalog(data)

    print(f"\nAll files written to:")
    print(f"  RAG docs: {RAG_DIR}")
    print(f"  Catalogs: {CATALOG_DIR}")
    print("\nDone.")


if __name__ == "__main__":
    main()
